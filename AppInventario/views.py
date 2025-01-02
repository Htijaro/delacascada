from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from AppInventario.forms import FormProducto
from .forms import FormProducto
from AppVentas.models import Domicilio
from .models import Producto
from openpyxl import Workbook
from datetime import datetime
from openpyxl.drawing.image import Image
from django.http import HttpResponse, JsonResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.views.decorators.cache import never_cache

# Se crean las vistas para el módulo de ventas. 

@never_cache
def dashinventario(request):
    productos = Producto.objects.all()
    return render(request, 'inventario/inventario.html', {'productos':productos})

def inventario_caducidad(request):
    productos = Producto.objects.all()
    return render(request, 'inventario/inventario_control_caducidad.html', {'productos':productos})

def inventario_analisis(request):
    productos = Producto.objects.all()
    return render(request, 'inventario/inventario_analisis.html', {'productos':productos})

def inventario_devoluciones(request):
    productos = Producto.objects.all()
    return render(request, 'inventario/inventario_devoluciones.html', {'productos':productos})

def inventario_productos(request):
    productos = Producto.objects.all()
    productoss = Producto.objects.all()
    return render(request, 'inventario/inventario_productos.html', {'productoss': productoss,'productos':productos})


# Productos:


def exportar_excel_productos(request):
    # Crear un nuevo libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'

    # Incluir la imagen
    img = Image('static/images/LogoCASWhite.png')
    img.anchor = 'B1'
    img.width = int(img.width * 0.1)
    img.height = int(img.height * 0.1)
    ws.add_image(img)

    # Crear el título en la hoja
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B1'].fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
    ws['B1'].font = Font(name='Calibri', size=18, color='FFFFFFFF', bold=True)
    ws['B1'] = 'REPORTE DE \n Productos'

        
    # Incluir la fecha del reporte
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['B2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B2'].font = Font(name='Calibri', size=12, color='003459', bold=True)  
    ws['B2'] = f'Fecha de creación: {current_date}'

    # Cambiar características de las celdas
    ws.merge_cells('B1:F1')
    ws.row_dimensions[1].height = 39
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
 

    

    # Crear la cabecera
    headers = ['Nombre Producto', 'Cantidad Producto', 'Precio Producto', 'Fecha Actualización', 
               'Otros Datos']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=3, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell.fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
        cell.font = Font(name='Calibri', size=12, color='FFFFFFFF', bold=True)
        cell.value = header

    # Obtener los productos de la base de datos
    productos = Producto.objects.all()
    for row_num, producto in enumerate(productos, 4):
        ws.cell(row=row_num, column=2).value = producto.nombreproductoinv
        ws.cell(row=row_num, column=3).value = producto.cantidadproducto
        ws.cell(row=row_num, column=4).value = producto.precioproductoinv
        ws.cell(row=row_num, column=5).value = producto.fechaactualizacioninv
        ws.cell(row=row_num, column=6).value = producto.otrosdatos


    # Establecer el nombre del archivo
    nombre_archivo = "Reporte De La Cascada Productos.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response

#Productos:

@login_required
def form_productos(request):
    if request.method == 'POST':
        form = FormProducto(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save(commit=False)
            nombreproductoinv = form.cleaned_data['nombreproductoinv']
            cantidadproducto = form.cleaned_data['cantidadproducto']
            precioproductoinv = form.cleaned_data['precioproductoinv']
            fechaactualizacioninv = form.cleaned_data['fechaactualizacioninv']
            otrosdatos = form.cleaned_data['otrosdatos']
            presentacion = form.cleaned_data['presentacion']
            foto_producto = form.cleaned_data['foto_producto']
            
            nuevo_producto = Producto.objects.create(
                nombreproductoinv=nombreproductoinv,
                cantidadproducto=cantidadproducto,
                precioproductoinv=precioproductoinv,
                fechaactualizacioninv=fechaactualizacioninv,
                otrosdatos=otrosdatos,
                presentacion=presentacion,
                foto_producto=foto_producto,
            )

            return redirect('inventario_productos')
    else:
        form = FormProducto()
    
    data = {
        'formproducto': form,
    }
    
    return render(request, 'inventario/FormProducto.html', data)


def crear_producto(request):
    if request.method == 'POST':
        form = FormProducto(request.POST, request.FILES)
        if form.is_valid():
            # Depuración: imprime los datos válidos del formulario
            print('Datos válidos:', form.cleaned_data)
            form.save()
            return redirect('AppInventario/InventarioProductos')  # Cambia a la vista que quieras redirigir después de guardar
        else:
            # Depuración: imprime los errores del formulario
            print('Errores del formulario:', form.errors)
    else:
        form = FormProducto()
    
    return render(request, 'formproducto', {'formproducto': form})


@login_required
def editar_productos(request, id):
    producto = get_object_or_404(Producto, idproducto=id)

    if request.method == 'POST':
        formulario = FormProducto(request.POST, instance=producto)
        if formulario.is_valid():
            formulario.save()
            return redirect('inventario_productos')
    else:
        formulario = FormProducto(instance=producto)

    return render(request, 'inventario/modificarProductos.html', {'producto': producto})
    
def formulario_productos(request):
    if request.method == 'POST':
        form = FormProducto(request.POST)
        if form.is_valid():
            id_domicilio = form.cleaned_data['iddomicilio']
            domicilio = Domicilio.objects.filter(iddomicilio=id_domicilio).first()
            if domicilio:
                producto = form.save(commit=False)
                producto.iddomicilio = domicilio
                producto.save()
                return redirect('ruta_exitosa')
            else:
                form.add_error('iddomicilio', 'El domicilio especificado no existe.')
    else:
        form = FormProducto()

    return render(request, 'inventario/formulario_productos.html', {'form': form})
    
@login_required
def eliminar_Productos(request, id):
    productos = get_object_or_404(Producto, idproducto=id)
    productos.delete()
    
    return redirect(to="inventario_productos")
