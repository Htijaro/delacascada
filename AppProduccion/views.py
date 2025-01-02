from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from .models import Fichaproduccion, Controlcalidad 
from .forms import  FormControlcalidad, FormFichaproduccion
from AppInventario.models import Producto
from .resources import ReporteFichaproduccion,ReporteControlcalidad
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from datetime import datetime
from pyexpat.errors import messages
from django.views.decorators.cache import never_cache

# Se crean las vistas para el módulo de ventas. 

@never_cache
@login_required  
def dashproduccion(request):
    fichaproduccions = Fichaproduccion.objects.all()
    productos = Producto.objects.all()
    return render(request, 'produccion/produccion.html', {'fichaproduccions': fichaproduccions, 'productos':productos})


@login_required  
def produccion_calidad(request):
    fichaproduccions = Fichaproduccion.objects.all()
    productos = Producto.objects.all()
    return render(request, 'produccion/produccion_calidad.html', {'fichaproduccions': fichaproduccions, 'productos':productos})


@login_required  
def produccion_mantenimiento(request):
    fichaproduccions = Fichaproduccion.objects.all()
    productos = Producto.objects.all()
    return render(request, 'produccion/produccion_mantenimiento.html', {'fichaproduccions': fichaproduccions, 'productos':productos})


@login_required  
def produccion_materiasP(request):
    fichaproduccions = Fichaproduccion.objects.all()
    productos = Producto.objects.all()
    return render(request, 'produccion/produccion_materias_prima.html', {'fichaproduccions': fichaproduccions, 'productos':productos})


#REPORTES

def exportar_excel_produccion(request):
        # Crear un nuevo libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = 'Producción'

    # Incluir la imagen
    img = Image('static/images/LogoCASWhite.png')
    img.anchor = 'B1'
    img.width = int(img.width * 0.1)
    img.height = int(img.height * 0.1)
    ws.add_image(img)

    # Crear el título en la hoja
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B1'].fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
    ws['B1'].font = Font(name='Calibri', size=18, color='FFFFFFFF', bold=True)
    ws['B1'] = 'REPORTE DE \n PRODUCCIÓN'

        
    # Incluir la fecha del reporte
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['B2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B2'].font = Font(name='Calibri', size=12, color='003459', bold=True)  
    ws['B2'] = f'Fecha de creación: {current_date}'

    # Cambiar características de las celdas
    ws.merge_cells('B1:L1')
    ws.row_dimensions[1].height = 39
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    

    # Crear la cabecera
    headers = ['Id Produccion', 'Cantidad Produccion', 'Medición Clorado', 'Fecha Producción', 'Hora Produccion', 'Medición PH', 
               'Filtrado', 'Microfiltrado', 'Empaque', 'Número lote', 'Estado produccion']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=3, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell.fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
        cell.font = Font(name='Calibri', size=12, color='FFFFFFFF', bold=True)
        cell.value = header

    # Obtener los pedidos de la base de datos
    produccion = Fichaproduccion.objects.all()
    for row_num, produccion in enumerate(produccion, 4):
        ws.cell(row=row_num, column=2).value = produccion.idfichaproduccion
        ws.cell(row=row_num, column=3).value = produccion.cantidadproduccion
        ws.cell(row=row_num, column=4).value = produccion.medirclorado
        ws.cell(row=row_num, column=5).value = produccion.fecha_produccion
        ws.cell(row=row_num, column=6).value = produccion.hora_produccion
        ws.cell(row=row_num, column=7).value = produccion.medirph
        ws.cell(row=row_num, column=8).value = produccion.filtrado
        ws.cell(row=row_num, column=9).value = produccion.microfiltrado
        ws.cell(row=row_num, column=10).value = produccion.empaque
        ws.cell(row=row_num, column=11).value = produccion.num_lote_produccion 
        ws.cell(row=row_num, column=12).value = produccion.estadoproduccion

    # Establecer el nombre del archivo
    nombre_archivo = "Reporte De La Cascada Producción.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response

def exportar_excel_Controlcalidad(request):
    reporte_resource = ReporteControlcalidad()
    dataset = reporte_resource.export()

    response = HttpResponse(dataset.export('xls'), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="controlcalidad.xls"'
    
    return response

#FORMULARIOS
# Ficha producción:
        
@login_required
def fichaproduccion(request):
    if request.method == 'POST':
        formulario = FormFichaproduccion(request.POST)
        if formulario.is_valid():
            produccion = formulario.save(commit=False)
            idfichaproduccion = formulario.cleaned_data['idfichaproduccion']
            cantidadproduccion = formulario.cleaned_data['cantidadproduccion']
            medirclorado = formulario.cleaned_data['medirclorado']
            medirph = formulario.cleaned_data['medirph']
            filtrado = formulario.cleaned_data['filtrado']
            microfiltrado = formulario.cleaned_data['microfiltrado']
            empaque = formulario.cleaned_data['empaque']
            num_lote_produccion = formulario.cleaned_data['num_lote_produccion']
            estadoproduccion = formulario.cleaned_data['estadoproduccion']
            
            nuevo_pedido = Fichaproduccion.objects.create(
                idfichaproduccion=idfichaproduccion,
                cantidadproduccion=cantidadproduccion,
                medirclorado=medirclorado,
                medirph=medirph,
                filtrado=filtrado,
                microfiltrado=microfiltrado,
                empaque=empaque,
                num_lote_produccion=num_lote_produccion,
                estadoproduccion=estadoproduccion,
            )

            return redirect('dashproduccion')
    else:
        formulario = FormFichaproduccion()
    
    data = {
        'formfichaproduccion': formulario,
    }
    
    return render(request, 'produccion/FormFichaproduccion.html', data)



@login_required
def editarFichaproduccion(request, id):
    fichaproduccion = get_object_or_404(Fichaproduccion, idfichaproduccion=id)

    if request.method == 'POST':
        formulario = FormFichaproduccion(request.POST, instance=fichaproduccion)
        if formulario.is_valid():
            formulario.save()
            return redirect('dasproduccion')
    else:
        formulario = FormFichaproduccion(instance=fichaproduccion)

    return render(request, 'produccion/modificarProduccion.html', {'formmodificarproduccion': formulario})


@login_required
def eliminarFichaproduccion(request, id):
    produccion = get_object_or_404(Fichaproduccion, idproduccion=id)
    produccion.delete()
    
    return redirect(to="dashproduccion")



# Control Calidad:


@login_required
def controlcalidad(request):
    data = {
        'formcontrolcalidad':FormControlcalidad()
    }
    
    if request.method == 'POST':
        formulario = FormControlcalidad(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            data["mensaje"] = "Control Calidad guardado exitosamente"
        else:
            data["formcontrolcalidad"] = formulario
    
    return render(request, 'produccion/FormControlcalidad.html', data )


@login_required  
def dashproduccion_controlcalidad(request):
    formcontrolcalidads = FormControlcalidad.objects.all()
    return render(request, 'produccion.html', {'formcontrolcalidads': formcontrolcalidads})
    
    
@login_required
def editarControlcalidad(request, id):
    controlcalidad = get_object_or_404(Controlcalidad, idcontrolcalidad=id)
    
    data = {
        'formmodificarControlCalidad': FormControlcalidad(instance=controlcalidad)
    }
    
    if request.method == 'POST':
        formulario = FormControlcalidad(data=request.POST, instance=controlcalidad, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            return redirect(to="dashproduccion")  
        else:
            data["formmodificarControlCalidad"] = formulario  
    
    return render(request, 'produccion/modificarControlCalidad.html', data)
    
    
@login_required
def eliminarControlcalidad(request, id):
    controlcalidad = get_object_or_404(Controlcalidad, idcontrolcalidad=id)
    controlcalidad.delete()
    
    return redirect(to="dashproduccion")


def produccion(request):
    return render(request, 'produccion/produccion.html')
  
  

"""    
@login_required
def fichaproduccion(request):
    data = {

        'formfichaproduccion':FormFichaproduccion()
    }
    
    if request.method == 'POST':
        formulario = FormFichaproduccion(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            data["mensaje"] = "Producción guardada exitosamente"
        else:
            data["formfichaproduccion"] = formulario
    
    return render(request, 'FormFichaproduccion.html', data, )


@login_required  
def dashproduccion(request):
    fichaproduccions = Fichaproduccion.objects.all()
    return render(request, 'produccion.html', {'fichaproduccions': fichaproduccions})
    
    
@login_required
def editarFichaproduccion(request, id):
    fichaproduccion = get_object_or_404(Fichaproduccion, idfichaproduccion=id)
    
    data = {
        'formmodificarfichaproduccion': FormFichaproduccion(instance=fichaproduccion)
    }
    
    if request.method == 'POST':
        formulario = FormFichaproduccion(data=request.POST, instance=fichaproduccion, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            return redirect(to="dashproduccion")  
        else:
            data["formmodificarfichaproduccion"] = formulario  
    
    return render(request, 'modificarProduccion.html', data)
    
    
@login_required
def eliminarFichaproduccion(request, id):
    fichaproduccion = get_object_or_404(Fichaproduccion, idfichaproduccion=id)
    fichaproduccion.delete()
    
    return redirect(to="dashproduccion")
"""



        

