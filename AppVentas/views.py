from pyexpat.errors import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from datetime import datetime
from AppInventario.models import Producto
from .models import Pedidos, Domicilio, Ventas
from .forms import FormDomicilio, FormPedido, FormVenta
from .models import Factura
from .forms import FacturaForm
from django.views.decorators.cache import never_cache

# Se crean las vistas para el módulo de ventas. 

@never_cache
@login_required  
def dashventas(request):
    pedidoss = Pedidos.objects.all()
    productos = Producto.objects.all()
    ventas = Ventas.objects.all()
    return render(request, 'ventas/ventas.html', {'pedidoss': pedidoss, 'productos':productos, 'ventas':ventas})

@login_required  
def ventas_pedidos(request):
    pedidoss = Pedidos.objects.all()
    productos = Producto.objects.all()
    return render(request, 'ventas/ventas_pedidos.html', {'pedidoss': pedidoss, 'productos':productos})

@login_required  
def ventas_clientes(request):
    pedidoss = Pedidos.objects.all()
    productos = Producto.objects.all()
    return render(request, 'ventas/ventas_clientes.html', {'pedidoss': pedidoss, 'productos':productos})

@login_required  
def ventas_productos(request):
    pedidoss = Pedidos.objects.all()
    productos = Producto.objects.all()
    return render(request, 'ventas/ventas_productos.html', {'pedidoss': pedidoss, 'productos':productos})

@login_required  
def ventas_domicilios(request):
    pedidoss = Pedidos.objects.all()
    productos = Producto.objects.all()
    domicilios = Domicilio.objects.all()
    return render(request, 'ventas/ventas_domicilios.html', {'pedidoss': pedidoss, 'productos':productos, 'domicilios':domicilios})


#REPORTES


#FORMULARIOS




#Pedidos:

@login_required
def form_pedidos(request):
    if request.method == 'POST':
        formulario = FormPedido(request.POST)
        if formulario.is_valid():
            pedido = formulario.save(commit=False)
            usuario_pedido = formulario.cleaned_data['usuario_pedido']
            direccionpedido = formulario.cleaned_data['direccionpedido']
            telefonopedido = formulario.cleaned_data['telefonopedido']
            estadopedido = formulario.cleaned_data['estadopedido']
            observacion = formulario.cleaned_data['observacion']
            producto = formulario.cleaned_data['producto']
            cantidad = formulario.cleaned_data['cantidad']
            precioproductoinv = formulario.cleaned_data['precioproductoinv']
            total = formulario.cleaned_data['total']
            
            nuevo_pedido = Pedidos.objects.create(
                usuario_pedido=usuario_pedido,
                direccionpedido=direccionpedido,
                telefonopedido=telefonopedido,
                estadopedido=estadopedido,
                observacion=observacion,
                producto=producto,
                cantidad=cantidad,
                precioproductoinv=precioproductoinv,
                total=total,
            )

            # Actualizar la cantidad del producto
            producto.cantidadproducto -= cantidad
            #producto.precioproducto -= producto
            #producto.total -= total
            producto.save()

            return redirect('ventas_pedidos')
    else:
        formulario = FormPedido()
    
    data = {
        'formpedido': formulario,
    }
    
    return render(request, 'ventas/FormPedido.html', data)



@login_required
def editarPedidos(request, id):
    pedido = get_object_or_404(Pedidos, idpedido=id)

    if request.method == 'POST':
        formulario = FormPedido(request.POST, instance=pedido)
        if formulario.is_valid():
            formulario.save()
            return redirect('ventas_pedidos')
    else:
        formulario = FormPedido(instance=pedido)

    return render(request, 'ventas/modificarPedidos.html', {'formmodificarpedido': formulario})
    
def formulario_pedidos(request):
    if request.method == 'POST':
        form = FormPedido(request.POST)
        if form.is_valid():
            id_domicilio = form.cleaned_data['iddomicilio']
            domicilio = Domicilio.objects.filter(iddomicilio=id_domicilio).first()
            if domicilio:
                pedido = form.save(commit=False)
                pedido.iddomicilio = domicilio
                pedido.save()
                return redirect('ruta_exitosa')
            else:
                form.add_error('iddomicilio', 'El domicilio especificado no existe.')
    else:
        form = FormPedido()

    return render(request, 'ventas/formulario_pedidos.html', {'form': form})
    
@login_required
def eliminarPedidos(request, id):
    pedidos = get_object_or_404(Pedidos, idpedido=id)
    pedidos.delete()
    
    return redirect(to="ventas_pedidos")


def exportar_excel_pedidos(request):
    # Crear un nuevo libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    # Incluir la imagen
    img = Image('static/images/LogoCASWhite.png')
    img.anchor = 'B1'
    img.width = int(img.width * 0.1)
    img.height = int(img.height * 0.1)
    ws.add_image(img)

    # Crear el título en la hoja
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B1'].fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
    ws['B1'].font = Font(name='Calibri', size=18, color='FFFFFFFF', bold=True)
    ws['B1'] = 'REPORTE DE \n PEDIDOS'

        
    # Incluir la fecha del reporte
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['B2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B2'].font = Font(name='Calibri', size=12, color='003459', bold=True)  
    ws['B2'] = f'Fecha de creación: {current_date}'

    # Cambiar características de las celdas
    ws.merge_cells('B1:L1')
    ws.row_dimensions[1].height = 39
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    

    # Crear la cabecera
    headers = ['Usuario Pedido', 'Dirección Pedido', 'Teléfono Pedido', 'Fecha Pedido', 'Hora Pedido', 'Estado Pedido', 
               'Observación', 'Producto', 'Cantidad', 'Precio Producto', 'Total']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=3, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell.fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
        cell.font = Font(name='Calibri', size=12, color='FFFFFFFF', bold=True)
        cell.value = header

    # Obtener los pedidos de la base de datos
    pedidos = Pedidos.objects.all()
    for row_num, pedido in enumerate(pedidos, 4):
        ws.cell(row=row_num, column=2).value = pedido.usuario_pedido
        ws.cell(row=row_num, column=3).value = pedido.direccionpedido
        ws.cell(row=row_num, column=4).value = pedido.telefonopedido
        ws.cell(row=row_num, column=5).value = pedido.fechapedido
        ws.cell(row=row_num, column=6).value = pedido.horapedido
        ws.cell(row=row_num, column=7).value = pedido.estadopedido
        ws.cell(row=row_num, column=8).value = pedido.observacion
        ws.cell(row=row_num, column=9).value = pedido.producto
        ws.cell(row=row_num, column=10).value = pedido.cantidad
        ws.cell(row=row_num, column=11).value = pedido.precioproductoinv 
        ws.cell(row=row_num, column=12).value = pedido.total

    # Establecer el nombre del archivo
    nombre_archivo = "Reporte De La Cascada Pedidos.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response





# Ventas:

@login_required
def form_ventas(request):
    if request.method == 'POST':
        formulario = FormVenta(request.POST)
        if formulario.is_valid():
            venta = formulario.save(commit=False)
            idfactura = formulario.cleaned_data['idfactura']
            nombre_usuario = formulario.cleaned_data['nombre_usuario']
            idpedido = formulario.cleaned_data['idpedido']
            
            nueva_venta = Ventas.objects.create(
                idfactura=idfactura,
                nombre_usuario=nombre_usuario,
                idpedido=idpedido,
            )


            return redirect('dashventas')
    else:
        formulario = FormVenta()
    
    data = {
        'formventa': formulario,
    }
    
    return render(request, 'ventas/FormVenta.html', data)



@login_required
def editarVentas(request, id):
    venta = get_object_or_404(Ventas, idventa=id)

    if request.method == 'POST':
        formulario = FormVenta(request.POST, instance=venta)
        if formulario.is_valid():
            formulario.save()
            return redirect('dashventas')
    else:
        formulario = FormVenta(instance=venta)

    return render(request, 'ventas/modificarVentas.html', {'formmodificarventa': formulario})
    

@login_required
def eliminarVentas(request, id):
    ventas = get_object_or_404(Ventas, idventa=id)
    ventas.delete()
    
    return redirect(to="dashventas")



def exportar_excel_ventas(request):
    # Crear un nuevo libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    # Incluir la imagen
    img = Image('static/images/LogoCASWhite.png')
    img.anchor = 'B1'
    img.width = int(img.width * 0.1)
    img.height = int(img.height * 0.1)
    ws.add_image(img)

    # Crear el título en la hoja
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B1'].fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
    ws['B1'].font = Font(name='Calibri', size=18, color='FFFFFFFF', bold=True)
    ws['B1'] = 'REPORTE DE \n VENTAS'

        
    # Incluir la fecha del reporte
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['B2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B2'].font = Font(name='Calibri', size=12, color='003459', bold=True)  
    ws['B2'] = f'Fecha de creación: {current_date}'

    # Cambiar características de las celdas
    ws.merge_cells('B1:F1')
    ws.row_dimensions[1].height = 39
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    

    # Crear la cabecera
    headers = ['Id Factura', 'Fecha Venta', 'Hora Venta', 'Nombre Usuario', 
               'Id Pedido']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=3, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell.fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
        cell.font = Font(name='Calibri', size=12, color='FFFFFFFF', bold=True)
        cell.value = header

    # Obtener los pedidos de la base de datos
    ventas = Ventas.objects.all()
    for row_num, ventas in enumerate(ventas, 4):
        ws.cell(row=row_num, column=2).value = ventas.idfactura
        ws.cell(row=row_num, column=3).value = ventas.fechaventa
        ws.cell(row=row_num, column=4).value = ventas.horaventa
        ws.cell(row=row_num, column=5).value = ventas.nombre_usuario
        


    # Establecer el nombre del archivo
    nombre_archivo = "Reporte De La Cascada Ventas.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response



#Domicilios:

@login_required
def form_domicilios(request):
    if request.method == 'POST':
        formulario = FormDomicilio(request.POST)
        if formulario.is_valid():
            domicilio = formulario.save(commit=False)
            estadodomicilio = formulario.cleaned_data['estadodomicilio']
            idpedido = formulario.cleaned_data['idpedido']
            
            nuevo_domicilio = Domicilio.objects.create(
                estadodomicilio=estadodomicilio,
                idpedido=idpedido,
            )


            return redirect('ventas_domicilios')
    else:
        formulario = FormDomicilio()
    
    data = {
        'formdomicilio': formulario,
    }
    
    return render(request, 'ventas/FormDomicilio.html', data)



@login_required
def editarDomicilios(request, id):
    domicilio = get_object_or_404(Domicilio, iddomicilio=id)

    if request.method == 'POST':
        formulario = FormDomicilio(request.POST, instance=domicilio)
        if formulario.is_valid():
            formulario.save()
            return redirect('ventas_domicilios')
    else:
        formulario = FormDomicilio(instance=domicilio)

    return render(request, 'ventas/modificarDomicilios.html', {'formmodificardomicilio': formulario})
    

@login_required
def eliminarDomicilios(request, id):
    domicilio = get_object_or_404(Domicilio, iddomicilio=id)
    domicilio.delete()
    
    return redirect(to="ventas_domicilios")








# Obtener productos y precio.


def precio_producto_api(request, producto_id):
    try:
        producto = Producto.objects.get(pk=producto_id)
        precio = producto.precioproductoinv
        return JsonResponse({'precio': precio})
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)

def get_product_price(request):
    product_id = request.GET.get('product_id')
    if product_id:
        try:
            product = Producto.objects.get(idproducto=product_id)
            return JsonResponse({'price': product.precioproductoinv})
        except Producto.DoesNotExist:
            return JsonResponse({'price': '0'})
    return JsonResponse({'price': '0'})



@login_required
def domicilios(request):
    data = {
        'formdomicilio': FormDomicilio()
    }
    
    if request.method == 'POST':
        formulario = FormDomicilio(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            data["mensaje"] = "Domicilio guardado exitosamente"
        else:
            data["formdomicilio"] = formulario
    
    return render(request, 'ventas/FormDomicilio.html', data)


def exportar_excel_domicilios(request):
    # Crear un nuevo libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = 'Domicilios'

    # Incluir la imagen
    img = Image('static/images/LogoCASWhite.png')
    img.anchor = 'B1'
    img.width = int(img.width * 0.1)
    img.height = int(img.height * 0.1)
    ws.add_image(img)

    # Crear el título en la hoja
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B1'].fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
    ws['B1'].font = Font(name='Calibri', size=18, color='FFFFFFFF', bold=True)
    ws['B1'] = 'REPORTE DE \n DOMICILIOS'

        
    # Incluir la fecha del reporte
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['B2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B2'].font = Font(name='Calibri', size=12, color='003459', bold=True)  
    ws['B2'] = f'Fecha de creación: {current_date}'

    # Cambiar características de las celdas
    ws.merge_cells('B1:D1')
    ws.row_dimensions[1].height = 39
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18


    

    # Crear la cabecera
    headers = ['Fecha Domicilio', 'Hora Domicilio', 'Estado Domicilio']
    
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=3, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell.fill = PatternFill(start_color='003459', end_color='003459', fill_type="solid")
        cell.font = Font(name='Calibri', size=12, color='FFFFFFFF', bold=True)
        cell.value = header

    # Obtener los pedidos de la base de datos
    ventas = Domicilio.objects.all()
    for row_num, domicilios in enumerate(ventas, 4):
        ws.cell(row=row_num, column=2).value = domicilios.fechadomicilio
        ws.cell(row=row_num, column=4).value = domicilios.horadomicilio
        ws.cell(row=row_num, column=3).value = domicilios.estadodomicilio
        


    # Establecer el nombre del archivo
    nombre_archivo = "Reporte De La Cascada Domicilios.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response

# Facturación
def crear_factura(request):
    if request.method == 'POST':
        form = FacturaForm(request.POST)
        if form.is_valid():
            factura = form.save()
            messages.success(request, f'Factura {factura.id_factura} creada exitosamente.')
            return redirect('listar_facturas')
    else:
        form = FacturaForm()
    return render(request, 'crear_factura.html', {'form': form})

def listar_facturas(request):
    facturas = Factura.objects.all()
    return render(request, 'listar_facturas.html', {'facturas': facturas})
